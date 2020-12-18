import pandas as pd
from openpyxl import Workbook,load_workbook
try:
    form_data = pd.read_excel("Club.xlsx")
except Exception as e:
    print("File Not Found")
    exit()
else:
    print("Data Loaded Successfully..")
    print("Please wait...")

for index,student in form_data.iterrows():
    clubs = student['Club']
    clubs = clubs.split(",")
    for club in clubs:
        club_name = club.split("(")[0]
        try:
            workbook = load_workbook(club_name + '.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
        finally:
            sheet.append(tuple(student))
            workbook.save(filename=club_name + '.xlsx')

print("File Generation done Successfully.")

